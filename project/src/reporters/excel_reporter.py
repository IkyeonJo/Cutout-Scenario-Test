'''
엑셀 리포트 생성기
'''

import os
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger('cutout_scenario.excel_reporter')

class ExcelReporter:
    """엑셀 리포트 생성기"""
    
    def __init__(self, config):
        self.config = config
        
    def create_report(self, filtered_scenarios):
        """필터링된 Test Case 목록을 엑셀 파일로 저장"""
        logger.info("엑셀 리포트 생성 시작")
        
        # 출력 경로 설정
        output_dir = self.config['simulation']['output']['output_directory']
        report_dir = os.path.join(output_dir, 'reports')
        os.makedirs(report_dir, exist_ok=True)
        
        excel_file = os.path.join(report_dir, self.config['simulation']['output']['excel_report_name'])
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Filtered Test Cases"
        
        # 헤더 설정
        headers = [
            "No.", 
            "Ego 속도 (km/h)", 
            "LV1 속도 (km/h)", 
            "LV1 횡방향 속도 (m/s)", 
            "초기 THW (s)", 
            "LV2 속도 (km/h)", 
            "LV2 감속도 (m/s²)", 
            "LV2 드러난 시점 THW (s)", 
            "TTC_reveal (s)", 
            "Human 모델 실패", 
            "ADS 모델 실패"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 데이터 입력
        for row_idx, scenario in enumerate(filtered_scenarios, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1)  # No.
            ws.cell(row=row_idx, column=2, value=scenario['v_ego'])
            ws.cell(row=row_idx, column=3, value=scenario['v_lv1'])
            ws.cell(row=row_idx, column=4, value=scenario['v_lv1_lat'])
            ws.cell(row=row_idx, column=5, value=scenario['thw_ego_lv1'])
            ws.cell(row=row_idx, column=6, value=scenario['v_lv2'])
            ws.cell(row=row_idx, column=7, value=scenario['dec_lv2'])
            ws.cell(row=row_idx, column=8, value=scenario['thw_ego_lv2_reveal'])
            ws.cell(row=row_idx, column=9, value=round(scenario['ttc_reveal'], 3) if scenario['ttc_reveal'] != float('inf') else 'inf')
            ws.cell(row=row_idx, column=10, value='Yes' if scenario['human_fails'] else 'No')
            ws.cell(row=row_idx, column=11, value='Yes' if scenario['ads_fails'] else 'No')
        
        # 열 너비 자동 조정
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # 요약 정보 시트 추가
        self._add_summary_sheet(wb, len(filtered_scenarios))
        
        # 파일 저장
        wb.save(excel_file)
        logger.info(f"엑셀 리포트 저장 완료: {excel_file}")
        
        return excel_file
    
    def _add_summary_sheet(self, workbook, filtered_count):
        """요약 정보 시트 추가"""
        ws = workbook.create_sheet(title="Summary")
        
        # 제목
        ws['A1'] = "Cut-out 시나리오 테스트 요약"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # 요약 정보
        ws['A3'] = "총 Concrete Scenario 수:"
        ws['B3'] = "계산 필요"  # 실제 값은 메인 스크립트에서 설정 필요
        
        ws['A4'] = "필터링된 Test Case 수:"
        ws['B4'] = filtered_count
        
        ws['A5'] = "시뮬레이션 대상 샘플 수:"
        ws['B5'] = self.config['simulation']['sampling']['target_sample_count']
        
        # UN R157 2023년 1월 개정안 적용 정보
        ws['A7'] = "적용 모델 기준:"
        ws['B7'] = "UN R157 2023년 1월 개정안"
        
        ws['A8'] = "Human 모델 반응 시간:"
        ws['B8'] = f"{self.config['control_models']['human_model']['reaction_time']}s"
        
        ws['A9'] = "Human 모델 최대 감속도:"
        ws['B9'] = f"{self.config['control_models']['human_model']['max_deceleration']}m/s²"
        
        ws['A10'] = "ADS 모델 반응 시간:"
        ws['B10'] = f"{self.config['control_models']['ads_model']['reaction_time']}s"
        
        ws['A11'] = "ADS 모델 최대 감속도:"
        ws['B11'] = f"{self.config['control_models']['ads_model']['max_deceleration']}m/s²"
        
        # 열 너비 설정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        return ws
    
    def add_sampled_scenarios(self, excel_file, sampled_scenarios):
        """샘플링된 최종 Test Case 목록을 별도 시트에 추가"""
        logger.info(f"엑셀 리포트에 샘플링 결과 추가: {excel_file}")
        
        # 워크북 로드
        wb = openpyxl.load_workbook(excel_file)
        
        # 새 시트 생성
        ws = wb.create_sheet(title="Sampled Test Cases")
        
        # 헤더 설정
        headers = [
            "No.", 
            "시나리오 ID",
            "Ego 속도 (km/h)", 
            "LV1 속도 (km/h)", 
            "LV1 횡방향 속도 (m/s)", 
            "초기 THW (s)", 
            "LV2 속도 (km/h)", 
            "LV2 감속도 (m/s²)", 
            "LV2 드러난 시점 THW (s)", 
            "TTC_reveal (s)"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 데이터 입력
        for row_idx, scenario in enumerate(sampled_scenarios, 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1)  # No.
            ws.cell(row=row_idx, column=2, value=f"scenario_{row_idx-1:03d}")
            ws.cell(row=row_idx, column=3, value=scenario['v_ego'])
            ws.cell(row=row_idx, column=4, value=scenario['v_lv1'])
            ws.cell(row=row_idx, column=5, value=scenario['v_lv1_lat'])
            ws.cell(row=row_idx, column=6, value=scenario['thw_ego_lv1'])
            ws.cell(row=row_idx, column=7, value=scenario['v_lv2'])
            ws.cell(row=row_idx, column=8, value=scenario['dec_lv2'])
            ws.cell(row=row_idx, column=9, value=scenario['thw_ego_lv2_reveal'])
            ws.cell(row=row_idx, column=10, value=round(scenario['ttc_reveal'], 3) if scenario['ttc_reveal'] != float('inf') else 'inf')
        
        # 열 너비 자동 조정
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # 요약 정보 시트 업데이트
        if 'Summary' in wb.sheetnames:
            summary_ws = wb['Summary']
            summary_ws['B5'] = len(sampled_scenarios)
        
        # 파일 저장
        wb.save(excel_file)
        logger.info(f"샘플링 결과 추가 완료: {excel_file}")
        
        return excel_file
